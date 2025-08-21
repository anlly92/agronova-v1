import traceback
from django.shortcuts import render,redirect,get_object_or_404 #obtiene un objeto del modelo, o devuelve un error 404 si no se encuentra.
from django.db.models import Sum, F, Q # sumar, referenciar campos del modelo, clase que sirve para construir consultas complejas
from .forms import LoteForm, RecoleccionForm, PagosForm # formularios importados desde la  misma app
from django.contrib import messages #mensajes de exito y error
from .models import Lote, Recoleccion, Empleado, Pagos # modelos utilizados en las consultas
from decimal import Decimal, InvalidOperation ##excepcion que se lanza si hay error al convertir un string a decimal
from inventarios.utils import normalizar_texto, es_numero, parsear_fecha # funciones que se encunetran en utils en la app de inventarios

from django.http import HttpResponse # devuelve una respuesta http personalizada como un archivo excel o pdf descaegable 

import openpyxl # para generar documentos excel
from django.template.loader import get_template # carga una plantilla html para convertir a pdf o excel
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side # estilos para exce, funete, alineacion de texto, relleno, bordes
from xhtml2pdf import pisa # para generar pdf desde html y css
from django.db.models.functions import ExtractYear, ExtractMonth # para extraer partes de fechas en concultas 
from validaciones import validar_campos_especificos

from administracion.decorators import solo_admin_principal # lo importamos para manejar la proteccion de las rutas si no es administrador principal
from django.contrib.auth.decorators import login_required


@login_required
def obtener_datos_recoleccion_filtrados(request):
    año = request.GET.get("año")
    mes = request.GET.get("mes")
    lote = request.GET.get("lote", "").strip().lower() # bloque que toma los parametros enviados desde el formulario Get

    recolecciones = Recoleccion.objects.all() # obtiene todsos los registros del modelo

    # Solo filtrar si año y mes son números válidos y numericos 
    if año and año.isdigit():
        recolecciones = recolecciones.filter(fecha__year=int(año))
    if mes and mes.isdigit():
        recolecciones = recolecciones.filter(fecha__month=int(mes))
    if lote:
        recolecciones = recolecciones.filter(id_lote_nombre_icontains=lote)

    consulta_total = (
        recolecciones
        .annotate(
            año=ExtractYear('fecha'),
            mes=ExtractMonth('fecha'),
            nombre_lote_anotado=F('id_lote__nombre')
        )
        .values('año', 'mes', 'nombre_lote_anotado')
        .annotate(total_kilos=Sum('kilos'))
        .order_by('-año', '-mes')
    )# bloque que agrupa y resume la informacion por año, mes y lote, calcula el total de kilos y devuelve una lista de diccionarios 
    return consulta_total 

@login_required
def exportar_total_recolecciones_excel(request):
    datos = obtener_datos_recoleccion_filtrados(request)  # usa los filtros del GET,es decir obtiene los datos filtrados segun la solicitud

    wb = openpyxl.Workbook() # crea una nueva hora de excel
    ws = wb.active
    ws.title = "Total Recolección" # tituilo de la hora de excel

    # Estilos
    titulo_font = Font(bold=True, size=14) # estilo titulo dle informe
    encabezado_font = Font(bold=True, color='FFFFFF') # asegura que el texto sea visible en un fondo blanco 
    encabezado_fill = PatternFill(start_color='3D7A1F', end_color='3D7A1F', fill_type='solid') #  Sirve para resaltar visualmente los encabezados de la tabla.

    alineacion_centrada = Alignment(horizontal='center', vertical='center') # alinea el contenido de la celdas, horizontal y verticalmente 
    borde = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    ) # define borde para todas las celdas, para la representacion de tipo tabla

    fila_inicio = 15 # inicio desde donde se escriben los datos 
    columna_inicio = 5 # inicio 

    ws.merge_cells(start_row=fila_inicio, start_column=columna_inicio, end_row=fila_inicio, end_column=columna_inicio + 3) #une varias celdas en una sola para que el titulo ocupe varias columnas
    celda_titulo = ws.cell(row=fila_inicio, column=columna_inicio) #seleccional aprimera celda del rango combinado para poner el titulo
    celda_titulo.value = "Informe de Recolecciones" # se establece el titulo que aparecera en al hoja 
    celda_titulo.font = titulo_font # aplica el estilo que se definio antes
    celda_titulo.alignment = alineacion_centrada # centra horizontalmente l titulo dentro d ela ceda combinada

    encabezados = ['Año', 'Mes', 'Nombre del lote', 'Total de kilos recolectados']
    for i, encabezado in enumerate(encabezados): # itera sobre la lista de encabezdos 
        col = columna_inicio + i# calcula la columna de inicio de donde se colocan los enacabezados (ejm:5,6,7)
        celda = ws.cell(row=fila_inicio + 2, column=col, value=encabezado)#Crea una celda en la fila donde van los encabezados, dos filas debajo del título principal y pone el texto, año, mes 
        celda.font = encabezado_font #aplica estilos definidos anteriormete
        celda.fill = encabezado_fill
        celda.alignment = alineacion_centrada
        celda.border = borde

    for index, fila in enumerate(datos):# lista de diccionarios obtenidos
        fila_actual = fila_inicio + 3 + index # espacio entre el titulo, los encabezados y los registros y se suma idex para avanzar a ala siguiente fila en acda iteracion
        ws.cell(row=fila_actual, column=columna_inicio, value=fila["año"])  # pone los valores en cada columna segun corresponda el dato 
        ws.cell(row=fila_actual, column=columna_inicio + 1, value=fila["mes"])
        ws.cell(row=fila_actual, column=columna_inicio + 2, value=fila["nombre_lote"])
        ws.cell(row=fila_actual, column=columna_inicio + 3, value=round(fila["total_kilos"] or 0))

        for col in range(columna_inicio, columna_inicio + 4):# recorre 4 columnasque comiences desde inicio 
            celda = ws.cell(row=fila_actual, column=col) # obtiene la celda en la fila actual y en la columna
            celda.alignment = alineacion_centrada # alinea el cntenido
            celda.border = borde #aplica borde alrededor de la celda 

    for col_letra in ['E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letra].width = 22 # ajusta el ancho de las columnas 

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') # prepara la respuesta http cpon excel y  crea la respuesta  con el tipo de contenido propio de archivos excel 
    response['Content-Disposition'] = 'attachment; filename=informe_recolecciones.xlsx' # indica al navegador que debe descargar el archivo no abrirlo en el navegador 
    wb.save(response)#guarda el libro excel directamente en el response, es decir lo escribe como si fuer aun archivo
    return response # devuelve el archivo como respuesta 

@login_required
def exportar_total_recolecciones_pdf(request):
    datos = obtener_datos_recoleccion_filtrados(request)  # se llama a al funcion que filtra los datos que estan almeacenados 

    template_path = 'cafe_cardamomo/pdf_total_recoleccion.html' #Especifica qué plantilla HTML usar para generar el contenido del PDF.
    context = {'total_recoleccion': datos}#Crea un diccionario con los datos que serán enviados a la plantilla.
    
    response = HttpResponse(content_type='application/pdf') #Inicia una respuesta HTTP indicando que será un archivo PDF.
    response['Content-Disposition'] = 'attachment; filename="total_recoleccion.pdf"' #attachment fuerza la descarga del archivo y se define el nombre que vera el usuario 

    template = get_template(template_path) #Carga la plantilla y la renderiza con los datos enviados.
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)# Usa xhtml2pdf (Pisa) para convertir el HTML renderizado en PDF y se guarda directamente en la respuesta
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500) #Si hubo un error al generar el PDF, devuelve un mensaje con error 500.
    return response

@login_required
def total_de_recoleccion(request):
    consulta_total_recoleccion = obtener_datos_recoleccion_filtrados(request)
    
    año = request.GET.get("año")
    mes = request.GET.get("mes")
    lote = request.GET.get("lote", "").strip().lower()

    recolecciones = Recoleccion.objects.all()

    if año:
        recolecciones = recolecciones.filter(fecha__year=año)
    if mes:
        recolecciones = recolecciones.filter(fecha__month=mes)
    if lote:
        recolecciones = recolecciones.filter(nombre_lote__icontains=lote)

    # Agrupar por año, mes, nombre_lote y tipo_producto
    total_recoleccion = recolecciones.annotate(
        año=ExtractYear('fecha'),
        mes=ExtractMonth('fecha')
    ).values(
        'año',
        'mes',
        'nombre_lote',
        'tipo_producto'
    ).annotate(total_kilos=Sum('kilos')).order_by('nombre_lote', 'tipo_producto')

    cantidad_filas_vacias = 15 - total_recoleccion.count()

    contexto = {
        'total_recoleccion': total_recoleccion,
        'filas_vacias': range(cantidad_filas_vacias),
        'año': año,
        'mes': mes,
        'lote': lote,
    }
    return render(request, 'cafe_cardamomo/total_de_recoleccion.html', contexto)


#Vistas para la gestion de los lotes 
@login_required
def gestionar_lote(request):
    ok = False 
    # para acciones de editar o borrar 
    if request.method == "POST":

        seleccion = request.POST.get("elemento")
        accion = request.POST.get("accion")

        if seleccion:
            if "," in seleccion:
                ids = seleccion.split(",")  
            else:
                ids = [seleccion]
        else:
            ids = []

        ids = [int(x) for x in ids if x.strip().isdigit()]

        if accion == "borrar" and ids:
            ok = True
            Lote.objects.filter(pk__in=ids).delete()

        elif accion == "editar" and len(ids) == 1:
            ok = True
            return redirect("actualizar_lote", seleccion=ids[0])

    lotes, buscar, id_lote, nombre, hectareas, tipo_arbusto, estado = filtrar_lotes(request)
    cantidad_filas_vacias = 15 -lotes.count()

    contexto ={
        "Lotes": lotes,
        "filas_vacias": range(cantidad_filas_vacias),
        "buscar": buscar,
        "id_lote": id_lote,
        "nombre": nombre,
        "filtro_hectareas": hectareas,
        "filtro_tipo_arbusto": tipo_arbusto,
        "filtro_estado": estado,
        "ok": ok,
    }
    return render(request, 'cafe_cardamomo/mostrar_lotes.html',contexto)

#----Registro de lotes---
@login_required
def registrar_lote (request):
    ok = False 
    errores = {}

    if request.method == 'POST':
        datos = request.POST
        form = LoteForm(datos)

        errores = validar_campos_especificos(post_data=datos)

        for campo, mensaje in errores.items():
            if campo in form.fields:
                form.add_error(campo, mensaje)

        if form.is_valid():
            lote = form.save(commit=False) 
            lote.save()
            ok = True
            form = LoteForm() 
    else:
        form = LoteForm()

    return render (request, 'cafe_cardamomo/registro_lote.html', {'form': form,'ok':ok})



@login_required
def actualizar_lote (request,seleccion):
    ok = False
    lote = get_object_or_404(Lote, pk=seleccion)

    if request.method == 'POST':
        hectareas = request.POST.get("hectareas","").strip()
        tipo_arbusto = request.POST.get("tipo_arbusto","").strip()
        estado = request.POST.get("estado","").strip()

        # Validación: al menos un campo debe estar lleno
        if not hectareas and not tipo_arbusto and not estado:
            messages.error(request, "Debes ingresar al menos un dato para actualizar.")
            return redirect("actualizar_lote", seleccion=seleccion)

        if hectareas:
            lote.hectareas = hectareas

        if tipo_arbusto:
            lote.tipo_arbusto = tipo_arbusto

        if estado:
            lote.estado = estado
        lote.save()
        ok = True

    return render(request, 'cafe_cardamomo/actualizar_lote.html', {'lote': lote, 'ok': ok})


#funcion para la barra de busqueda o filtro en lotes
@login_required
def filtrar_lotes(request):
    buscar = request.GET.get("buscar", "").strip()
    id_lote = request.GET.get("id_lote", "").strip()
    nombre = request.GET.get("nombre", "").strip()#-----
    tipo_arbusto = request.GET.get("tipo_arbusto", "")
    estado = request.GET.get("estado", "")
    hectareas = request.GET.get("hectareas")

    lotes = Lote.objects.all()

    if buscar:
        buscar_normalizado = normalizar_texto(buscar)
        filtro_numerico = Q()
        filtro_texto = Q()
        if es_numero(buscar_normalizado):
            try:
                filtro_numerico = (
                    Q(id_lote=int(buscar_normalizado)) |
                    Q(hectareas=int(buscar_normalizado))
                )
            except ValueError:
                pass


        filtro_texto = (
            Q(nombre__iexact=buscar_normalizado)|
            Q(tipo_arbusto__iexact=buscar_normalizado)|
            Q(estado__iexact=buscar_normalizado)
        )

        lotes = Lote.objects.filter(
            filtro_numerico | filtro_texto
        )

    #campos utilizados en la modal del filtro
    if tipo_arbusto:
        lotes = lotes.filter(tipo_arbusto=tipo_arbusto)
    if estado:
        lotes = lotes.filter(estado=estado)
    if hectareas:
        lotes = lotes.filter(hectareas=hectareas)

    return lotes, buscar, id_lote, nombre, hectareas, tipo_arbusto, estado

@login_required
def gestionar_recoleccion (request):
    lotes, recolecciones, empleados, buscar, id_empleado, id_lote, tipo_producto, kilos, horas, fecha, tipo_pago = filtrar_recoleccion(request)  
    cantidad_filas_vacias = 15 - recolecciones.count()

    contexto = {
        'recolecciones': recolecciones,
        'filas_vacias': range(cantidad_filas_vacias),
        'buscar': buscar,
        'lotes': lotes,
        'filtro_id_empleado': id_empleado,
        'filtro_id_lote': id_lote,
        'filtro_tipo_producto': tipo_producto,
        'filtro_kilos': kilos,
        'filtro_horas': horas,
        'filtro_fecha': fecha,
        'filtro_tipo_pago': tipo_pago,
        'empleados': empleados,
    }

    return render (request, 'cafe_cardamomo/mostrar_recoleccion.html',contexto)

@login_required
def filtrar_recoleccion(request):
    buscar = request.GET.get("buscar", "").strip()
    id_empleado = request.GET.get("id_empleado", "")
    id_lote = request.GET.get("id_lote", "")
    tipo_producto = request.GET.get("tipo_producto", "")
    kilos = request.GET.get("kilos", "")
    horas = request.GET.get("horas_trabajadas", "")
    fecha = request.GET.get("fecha", "")
    tipo_pago = request.GET.get("tipo_pago", "")

    recolecciones = Recoleccion.objects.all()
    empleados = Empleado.objects.all()
    lotes = Lote.objects.all()

    if buscar:
        buscar_normalizado = normalizar_texto(buscar)
        filtro_numerico = Q()
        filtro_texto = Q()
        filtro_fecha = Q()

        # Filtra si es número exacto
        if es_numero(buscar_normalizado):
            try:
                filtro_numerico = (
                    Q(id_recoleccion=int(buscar_normalizado)) |
                    Q(id_empleado__id_empleado=int(buscar_normalizado)) |
                    Q(id_lote__id_lote=int(buscar_normalizado)) |
                    Q(kilos=int(buscar_normalizado)) |
                    Q(horas_trabajadas=Decimal(buscar_normalizado)) |
                    Q(tipo_pago__valor=int(buscar_normalizado))
                )
            except (ValueError, InvalidOperation):
                pass

        partes = buscar_normalizado.split()
        for parte in partes:
            filtro_texto &= (
            Q(id_empleado_nombre_icontains=parte) |
            Q(id_empleado_apellido_icontains=parte)
            )


        filtro_otros_datos = (
            Q(tipo_producto__iexact=buscar_normalizado) |
            Q(tipo_pago_tipo_pago_iexact=buscar_normalizado) |
            Q(id_lote_nombre_iexact=buscar_normalizado) 
        )
        # Filtro por fecha
        fecha_parseada = parsear_fecha(buscar)
        if fecha_parseada:
            filtro_fecha = Q(fecha=fecha_parseada)
        # Consulta combinada
        recolecciones = Recoleccion.objects.filter(
            filtro_numerico | filtro_texto | filtro_fecha | filtro_otros_datos
        )

    if id_empleado:
        try:
            id_empleado = int(id_empleado)
            recolecciones = recolecciones.filter(id_empleado=id_empleado)
        except ValueError:
            pass

    if id_lote:
        try:
            id_lote = int(id_lote)
            recolecciones = recolecciones.filter(id_lote=id_lote)
        except ValueError:
            pass

    if tipo_producto:
        recolecciones = recolecciones.filter(tipo_producto=tipo_producto)

    if kilos:
        try:
            kilos_decimal = Decimal(kilos)
            recolecciones = recolecciones.filter(kilos=kilos_decimal)
        except InvalidOperation:
            pass

    if horas:
        try:
            horas_decimal = Decimal(horas)
            recolecciones = recolecciones.filter(horas_trabajadas=horas_decimal)
        except InvalidOperation:
            pass

    if fecha:
        recolecciones = recolecciones.filter(fecha=fecha)

    if tipo_pago:
        recolecciones = recolecciones.filter(tipo_pago__tipo_pago=tipo_pago)

    return lotes, recolecciones, empleados, buscar, id_empleado, id_lote, tipo_producto, kilos, horas, fecha, tipo_pago  

@login_required
def registrar_recoleccion (request):
    ok = False 

    if request.method == 'POST':
        form = RecoleccionForm(request.POST)

        if form.is_valid():
            recoleccion = form.save(commit=False) 

            empleado = recoleccion.id_empleado
            lote = recoleccion.id_lote
            recoleccion.valor_pago = recoleccion.tipo_pago.valor

            if empleado:
                recoleccion.nombre_empleado = empleado.nombre + " " + empleado.apellido

            if lote:
                recoleccion.nombre_lote = lote.nombre

            recoleccion.save()
            
            ok = True
    else:
        form = RecoleccionForm()

    empleados = Empleado.objects.all()
    lotes = Lote.objects.all()
    return render (request, 'cafe_cardamomo/registro_recoleccion.html', {'form': form,'ok':ok, 'empleados': empleados, 'lotes': lotes})

@solo_admin_principal
def registrar_pago(request):
    ok = False
    if request.method == 'POST':
        tipo = request.POST.get('tipo_pago')
        valor = request.POST.get('valor')

        if tipo and valor:
            # Actualiza o crea el registro
            pago, creado = Pagos.objects.update_or_create(
                tipo_pago=tipo,
                defaults={'valor': valor}
            )
            
            ok = True

    # Siempre definimos form (si quieres usar un formulario para mostrar campos)
    from .forms import PagosForm
    form = PagosForm()

    return render(request, 'cafe_cardamomo/registrar_pago.html', {'form': form, 'ok': ok})