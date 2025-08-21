from django.shortcuts import render, redirect,get_object_or_404
from .forms import IngresosEgresosform,VentasForm
from .models import IngresosEgresos,Ventas
from administracion.models import Administrador
from inventarios.models import Inventario
from django.contrib.auth.models import User
from .forms import IngresosEgresosform
from .models import IngresosEgresos
from django.db.models import Sum
from django.db.models.functions import ExtractDay
from django.db.models.functions import ExtractMonth
from django.http import JsonResponse
from datetime import datetime
from django.utils import timezone
from calendar import monthrange, month_name
from django.db.models import Q
from decimal import InvalidOperation
from django.contrib import messages
from inventarios.utils import normalizar_texto, es_numero, parsear_fecha # funciones que se encunetran en utils en la app de inventarios
# Importaciones que se utilizan para realizar al descarga en excel
import openpyxl # Se utiliza para que permita la descarga de el informe en excel 
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side # Define los estilos para el excel ya que no soporta css
from django.http import HttpResponse
from validaciones import validar_campos_especificos
from django.template.loader import get_template
from xhtml2pdf import pisa

from django.contrib.auth.decorators import login_required

@login_required
def ingresos_egresos(request):
    if request.method == "POST":
        seleccion = request.POST.get("elemento")
        accion = request.POST.get("accion")
        if seleccion:
            if accion == "borrar":
                get_object_or_404(IngresosEgresos, pk=seleccion).delete()
                return redirect("ingresos_egresos")

    buscar, Ingresos_Egresos, documento_admin, tipo, fecha, monto = filtrar_ingresos_egresos(request)
    cantidad_filas_vacias = 15 - Ingresos_Egresos.count()

    contexto = {
        'Ingresos_Egresos': Ingresos_Egresos,
        'filas_vacias': range(cantidad_filas_vacias),
        'buscar': buscar,
        'filtro_documento_admin': documento_admin,
        'filtro_tipo': tipo,
        'filtro_fecha': fecha,
        'filtro_monto': monto,
    }

    return render(request, 'ingresos_egresos/mostrar_ingresos_egresos.html', contexto)

@login_required
def registro_ingresos_egresos (request):
    ok = False 
    errores = {}

    if request.method == 'POST':
        datos = request.POST
        form = IngresosEgresosform(datos)

        errores = validar_campos_especificos(post_data=datos)

        for campo, mensaje in errores.items():
                if campo in form.fields:
                    form.add_error(campo, mensaje)

        if form.is_valid():
            ingreso_egreso = form.save(commit=False) 
            ingreso_egreso.id_admin = Administrador.objects.get(correo=request.user.username)
            ingreso_egreso.save()
            ok = True 
    else:
        form = IngresosEgresosform()

    return render(request, 'ingresos_egresos/registro_ingresos_egresos.html', {
        'form': form,
        'ok':ok})

def filtrar_ingresos_egresos(request):
    documento_admin = request.GET.get("documento_admin", "")
    tipo = request.GET.get("tipo", "")
    fecha = request.GET.get("fecha", "")
    monto = request.GET.get("monto", "")
    buscar = request.GET.get("buscar", "").strip()

    Ingresos_Egresos = IngresosEgresos.objects.all()

    # Búsqueda general
    if buscar:
        buscar_normalizado = normalizar_texto(buscar)
        filtro_numerico = Q()
        filtro_texto = Q()
        filtro_fecha = Q()

        if es_numero(buscar_normalizado):
            try:
                filtro_numerico = (
                    Q(id_transaccion__iexact=buscar_normalizado) |
                    Q(id_admin__id_admin__iexact=buscar_normalizado) |
                    Q(monto=float(buscar_normalizado))
                )
            except (ValueError, InvalidOperation):
                filtro_numerico = Q()

        filtro_texto = (
            Q(tipo__iexact=buscar_normalizado) |
            Q(descripcion__icontains=buscar_normalizado) |
            Q(id_admin__nombre__iexact=buscar_normalizado)
        )

        fecha_parseada = parsear_fecha(buscar_normalizado)
        if fecha_parseada:
            filtro_fecha = Q(fecha=fecha_parseada)

        Ingresos_Egresos = Ingresos_Egresos.filter(
            filtro_numerico | filtro_texto | filtro_fecha
        )

    if documento_admin:
        try:
            documento_admin_int = int(documento_admin)
            Ingresos_Egresos = Ingresos_Egresos.filter(id_admin__id_admin=documento_admin_int)
        except ValueError:
            pass

    if tipo:
        Ingresos_Egresos = Ingresos_Egresos.filter(tipo__iexact=tipo)

    if fecha:
        Ingresos_Egresos = Ingresos_Egresos.filter(fecha=fecha)

    if monto:
        Ingresos_Egresos = Ingresos_Egresos.filter(monto=monto)

    return buscar, Ingresos_Egresos, documento_admin, tipo, fecha, monto

@login_required
def obtener_datos_ingresos_egresos_filtrados(request): # se define una función auxiliar para traer los datos de la funcion que filtra ingresos y egresos
    _, datos_filtrados, _, _, _, _ = filtrar_ingresos_egresos(request) # lo que hacen los guiones es ignorar lo que trae esta funcion y no necesita 
    return datos_filtrados # Retorna los datps filtrados

@login_required
def exportar_ingresos_egresos_excel(request):
    datos = obtener_datos_ingresos_egresos_filtrados(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ingresos y Egresos"

    # Estilos
    titulo_font = Font(bold=True, size=14)
    encabezado_font = Font(bold=True, color='FFFFFF')
    encabezado_fill = PatternFill(start_color='3D7A1F', end_color='3D7A1F', fill_type='solid')
    alineacion = Alignment(horizontal='center', vertical='center')
    borde = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    fila_inicio = 15
    columna_inicio = 3

    ws.merge_cells(start_row=fila_inicio, start_column=columna_inicio, end_row=fila_inicio, end_column=columna_inicio + 5)
    celda_titulo = ws.cell(row=fila_inicio, column=columna_inicio)
    celda_titulo.value = "Informe de Ingresos y Egresos"
    celda_titulo.font = titulo_font
    celda_titulo.alignment = alineacion

    encabezados = ['ID', 'Documento Admin', 'Nombre Admin', 'Tipo', 'Descripción', 'Fecha', 'Monto']
    for i, encabezado in enumerate(encabezados):
        col = columna_inicio + i
        celda = ws.cell(row=fila_inicio + 2, column=col, value=encabezado)
        celda.font = encabezado_font
        celda.fill = encabezado_fill
        celda.alignment = alineacion
        celda.border = borde

    for index, item in enumerate(datos):
        fila_actual = fila_inicio + 3 + index
        valores = [
            item.id_transaccion,
            item.id_admin.id_admin,
            item.id_admin.nombre,
            item.tipo,
            item.descripcion,
            item.fecha.strftime('%Y-%m-%d'),
            float(item.monto)
        ]

        for i, valor in enumerate(valores):
            celda = ws.cell(row=fila_actual, column=columna_inicio + i, value=valor)
            celda.alignment = alineacion
            celda.border = borde

    for col_letra in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col_letra].width = 22

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ingresos_egresos.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_ingresos_egresos_pdf(request):
    datos = obtener_datos_ingresos_egresos_filtrados(request)

    template_path = 'ingresos_egresos/pdf_ingresos_egresos.html'
    context = {'datos': datos}

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ingresos_egresos.pdf"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response

def ingresos_egresos(request):
    ok = False
    if request.method == "POST":

        seleccion = request.POST.get("elemento")
        accion = request.POST.get("accion")

        if seleccion:

            if "," in seleccion:
                ids = seleccion.split(",")  
            else:
                ids = [seleccion]
        else:
            ids = []

        ids = [int(x) for x in ids if x.strip().isdigit()]

        if accion == "borrar" and ids:
            ok = True
            for id_ingreso_egreso in ids:
                ingreso_egreso = get_object_or_404(IngresosEgresos, pk=id_ingreso_egreso)
                ingreso_egreso.anulada = True
                ingreso_egreso.save()
    
    buscar, Ingresos_Egresos, documento_admin, tipo, fecha, monto = filtrar_ingresos_egresos(request)
    Ingresos_Egresos = Ingresos_Egresos.filter(anulada=False) 
    cantidad_filas_vacias = 15 - Ingresos_Egresos.count()

    contexto = {
        'Ingresos_Egresos': Ingresos_Egresos,
        'ingresos_egresos': ingresos_egresos,
        'filas_vacias': range(cantidad_filas_vacias),
        'buscar': buscar,
        'filtro_documento_admin': documento_admin,
        'filtro_tipo': tipo,
        'filtro_fecha': fecha,
        'filtro_monto': monto,
        'ok': ok,
    }

    return render(request, 'ingresos_egresos/mostrar_ingresos_egresos.html', contexto)


def registro_ingresos_egresos (request):
    ok = False 
    if request.method == 'POST':
        form = IngresosEgresosform(request.POST)

        if form.is_valid():
            ingreso_egreso = form.save(commit=False) 
            ingreso_egreso.id_admin = Administrador.objects.get(correo=request.user.username)
            ingreso_egreso.save()

            ok = True 
    else:
        form = IngresosEgresosform()

    return render(request, 'ingresos_egresos/registro_ingresos_egresos.html', {'form': form,'ok':ok})


def filtrar_ingresos_egresos(request):
    documento_admin = request.GET.get("documento_admin", "")
    tipo = request.GET.get("tipo", "")
    fecha = request.GET.get("fecha", "")
    monto = request.GET.get("monto", "")
    buscar = request.GET.get("buscar", "").strip()

    Ingresos_Egresos = IngresosEgresos.objects.all()

    # Búsqueda general
    if buscar:
        buscar_normalizado = normalizar_texto(buscar)
        filtro_numerico = Q()
        filtro_texto = Q()
        filtro_fecha = Q()

        if es_numero(buscar_normalizado):
            try:
                filtro_numerico = (
                    Q(id_transaccion__iexact=buscar_normalizado) |
                    Q(id_admin__id_admin__iexact=buscar_normalizado) |
                    Q(monto=float(buscar_normalizado))
                )
            except (ValueError, InvalidOperation):
                filtro_numerico = Q()

        filtro_texto = (
            Q(tipo__iexact=buscar_normalizado) |
            Q(descripcion__icontains=buscar_normalizado) |
            Q(id_admin__nombre__iexact=buscar_normalizado)
        )

        fecha_parseada = parsear_fecha(buscar_normalizado)
        if fecha_parseada:
            filtro_fecha = Q(fecha=fecha_parseada)

        Ingresos_Egresos = Ingresos_Egresos.filter(
            filtro_numerico | filtro_texto | filtro_fecha
        )

    if documento_admin:
        try:
            documento_admin_int = int(documento_admin)
            Ingresos_Egresos = Ingresos_Egresos.filter(id_admin__id_admin=documento_admin_int)
        except ValueError:
            pass

    if tipo:
        Ingresos_Egresos = Ingresos_Egresos.filter(tipo__iexact=tipo)

    if fecha:
        Ingresos_Egresos = Ingresos_Egresos.filter(fecha=fecha)

    if monto:
        Ingresos_Egresos = Ingresos_Egresos.filter(monto=monto)

    return buscar, Ingresos_Egresos, documento_admin, tipo, fecha, monto


#-------------------------------------Informe anual----------------------------------------#
@login_required
def informe_anual(request):
    # Si no hay 'año', renderizar el HTML
    if 'anio' not in request.GET:
        return render(request, 'ingresos_egresos/informe_anual.html', {'now': datetime.now()})

    # Si hay 'año', procesar la solicitud JSON
    anio = request.GET.get('anio')
    if not anio:
        return JsonResponse({'error': 'Parámetro "anio" no proporcionado.'}, status=400)

    try:
        anio = int(anio)
    except ValueError:
        return JsonResponse({'error': 'Parámetro "anio" inválido. Debe ser un número entero.'}, status=400)

    # Validar rango de años
    current_year = datetime.now().year
    if anio < 1900 or anio > current_year:
        return JsonResponse({'error': f'El año debe estar entre 1900 y {current_year}.'}, status=400)

    # Consulta para ingresos
    ingresos = IngresosEgresos.objects.filter(fecha__year=anio, tipo='Ingreso').annotate(
        mes=ExtractMonth('fecha')
    ).values('mes').annotate(
        total=Sum('monto')
    ).order_by('mes')

    # Consulta para egresos
    egresos = IngresosEgresos.objects.filter(fecha__year=anio, tipo='Egreso').annotate(
        mes=ExtractMonth('fecha')
    ).values('mes').annotate(
        total=Sum('monto')
    ).order_by('mes')

    montos_ingresos = [0.0] * 12
    montos_egresos = [0.0] * 12

    for item in ingresos:
        mes_index = item['mes'] - 1
        montos_ingresos[mes_index] = float(item['total'] or 0.0)

    for item in egresos:
        mes_index = item['mes'] - 1
        montos_egresos[mes_index] = float(item['total'] or 0.0)

    etiquetas = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", 
                "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

    # Mensaje opcional si no hay datos
    if not ingresos and not egresos:
        return JsonResponse({
            'labels': etiquetas,
            'data': {
                'ingresos': montos_ingresos,
                'egresos': montos_egresos
            },
            'warning': f'No se encontraron datos para el año {anio}.'
        })

    return JsonResponse({
        'labels': etiquetas,
        'data': {
            'ingresos': montos_ingresos,
            'egresos': montos_egresos
        }
    })

@login_required
def descargar_informe_anual(request):
    anio = request.GET.get('anio')
    if not anio or not anio.isdigit() or len(anio) != 4:
        return HttpResponse('Año inválido', status=400)
    anio = int(anio)
    current_year = datetime.now().year
    if anio < 1900 or anio > current_year:
        return HttpResponse('Año fuera de rango', status=400)

    # Obtener datos
    ingresos = IngresosEgresos.objects.filter(fecha__year=anio, tipo='Ingreso')\
        .annotate(mes=ExtractMonth('fecha'))\
        .values('mes')\
        .annotate(total=Sum('monto'))\
        .order_by('mes')
    egresos = IngresosEgresos.objects.filter(fecha__year=anio, tipo='Egreso')\
        .annotate(mes=ExtractMonth('fecha'))\
        .values('mes')\
        .annotate(total=Sum('monto'))\
        .order_by('mes')

    ingresos_dict = {item['mes']: float(item['total'] or 0.0) for item in ingresos}
    egresos_dict = {item['mes']: float(item['total'] or 0.0) for item in egresos}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Informe {anio}"

    # Estilos
    bold_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    fill_header = PatternFill(start_color='4b7510', end_color='4b7510', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:D1')
    ws['A1'] = f"Informe Anual de Ingresos y Egresos - {anio}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center

    ws.append(['Mes', 'Ingresos', 'Egresos', 'Total'])
    for cell in ws[2]:
        cell.font = bold_font
        cell.alignment = center
        cell.fill = fill_header
        cell.border = border

    from calendar import month_name
    meses = [month_name[i] for i in range(1, 13)]
    for i, mes in enumerate(meses, 1):
        ingreso = ingresos_dict.get(i, 0.0)
        egreso = egresos_dict.get(i, 0.0)
        total = ingreso - egreso
        ws.append([mes, ingreso, egreso, total])
        for cell in ws[ws.max_row]:
            cell.alignment = center
            cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="informe_anual_{anio}.xlsx"'
    wb.save(response)
    return response


#----------------------------------Informe mensual -----------------------------------
@login_required
def informe_mensual(request):
    return render(request, 'ingresos_egresos/informe_mensual.html')

@login_required
# Vista para los datos del informe mensual
def datos_informe_mensual(request):
    # Obtener parámetros mes y anio
    mes_str = request.GET.get('mes')
    anio_str = request.GET.get('anio')
    
    # Validar que los parámetros estén presentes
    if not mes_str or not anio_str:
        return JsonResponse({'error': 'Parámetros requeridos'}, status=400)
    
    # Validar que los parámetros sean números
    try:
        mes = int(mes_str)
        anio = int(anio_str)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Valores inválidos'}, status=400)
    
    # Validar que el mes esté entre 1 y 12
    if mes < 1 or mes > 12:
        return JsonResponse({'error': 'Mes inválido'}, status=400)
    
    # Validar que el año esté en un rango razonable
    current_year = datetime.now().year
    if anio < 1900 or anio > current_year:
        return JsonResponse({'error': f'El año debe estar entre 1900 y {current_year}'}, status=400)
    
    # Obtener el número de días en el mes
    num_dias = monthrange(anio, mes)[1]
    
    # Obtener los datos de ingresos y egresos agrupados por día
    ingresos = IngresosEgresos.objects.filter(
        fecha__year=anio, fecha__month=mes, tipo='Ingreso'
    ).annotate(dia=ExtractDay('fecha')).values('dia').annotate(total=Sum('monto')).order_by('dia')
    
    egresos = IngresosEgresos.objects.filter(
        fecha__year=anio, fecha__month=mes, tipo='Egreso'
    ).annotate(dia=ExtractDay('fecha')).values('dia').annotate(total=Sum('monto')).order_by('dia')
    
    # Crear listas para los montos de ingresos y egresos
    montos_ingresos = [0.0] * num_dias
    montos_egresos = [0.0] * num_dias
    
    # Llenar las listas con los datos obtenidos
    for ingreso in ingresos:
        montos_ingresos[ingreso['dia'] - 1] = float(ingreso['total'])
    
    for egreso in egresos:
        montos_egresos[egreso['dia'] - 1] = float(egreso['total'])
    
    # Etiquetas para los días
    labels = [str(i) for i in range(1, num_dias + 1)]
    
    # Agregado manejo de datos vacíos con mensaje warning
    if not ingresos and not egresos:
        return JsonResponse({
            'warning': f'No se encontraron datos para el mes {mes}/{anio}',
            'labels': labels,
            'data': {'ingresos': montos_ingresos, 'egresos': montos_egresos}
        })
    
    # Retornar los datos en formato JSON
    return JsonResponse({
        'labels': labels,
        'data': {
            'ingresos': montos_ingresos,
            'egresos': montos_egresos
        }
    })


@login_required
def descargar_informe_mensual_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from .models import IngresosEgresos
    from django.db.models import Sum
    from django.db.models.functions import ExtractDay
    from calendar import monthrange
    # Obtener parámetros mes y anio
    mes_str = request.GET.get('mes')
    anio_str = request.GET.get('anio')
    if not mes_str or not anio_str:
        return HttpResponse('Parámetros mes y año requeridos', status=400)
    try:
        mes = int(mes_str)
        anio = int(anio_str)
    except (ValueError, TypeError):
        return HttpResponse('Mes o año inválido', status=400)
    if mes < 1 or mes > 12:
        return HttpResponse('Mes fuera de rango', status=400)
    current_year = datetime.now().year
    if anio < 1900 or anio > current_year:
        return HttpResponse('Año fuera de rango', status=400)
    # Obtener los datos de ingresos y egresos agrupados por día
    ingresos = IngresosEgresos.objects.filter(
        fecha__year=anio, fecha__month=mes, tipo='Ingreso'
    ).annotate(dia=ExtractDay('fecha')).values('dia').annotate(total=Sum('monto')).order_by('dia')
    egresos = IngresosEgresos.objects.filter(
        fecha__year=anio, fecha__month=mes, tipo='Egreso'
    ).annotate(dia=ExtractDay('fecha')).values('dia').annotate(total=Sum('monto')).order_by('dia')
    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Informe Mensual"
    # Estilos y fila de inicio
    encabezado_font = Font(bold=True, color='FFFFFF')
    encabezado_fill = PatternFill(start_color='3D7A1F', end_color='3D7A1F', fill_type='solid')
    alineacion = Alignment(horizontal='center', vertical='center')
    borde = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    fila_inicio = 2
    # Encabezados
    encabezados = ['Día', 'Ingresos', 'Egresos', 'Total']
    for i, encabezado in enumerate(encabezados, start=2):
        celda = ws.cell(row=fila_inicio + 2, column=i)
        celda.value = encabezado
        celda.font = encabezado_font
        celda.fill = encabezado_fill
        celda.alignment = alineacion
        celda.border = borde

    # Mapear ingresos y egresos por día
    num_dias = monthrange(anio, mes)[1]
    ingresos_por_dia = {item['dia']: item['total'] for item in ingresos}
    egresos_por_dia = {item['dia']: item['total'] for item in egresos}
    for dia in range(1, num_dias + 1):
        fila = fila_inicio + 2 + dia
        ingreso = float(ingresos_por_dia.get(dia, 0))
        egreso = float(egresos_por_dia.get(dia, 0))
        total = ingreso - egreso
        ws.cell(row=fila, column=2).value = dia
        ws.cell(row=fila, column=3).value = ingreso
        ws.cell(row=fila, column=4).value = egreso
        ws.cell(row=fila, column=5).value = total
        for col in range(2, 6):
            celda = ws.cell(row=fila, column=col)
            celda.alignment = alineacion
            celda.border = borde
    # Ajustar ancho de columnas
    for col in range(2, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=informe_mensual_{mes}_{anio}.xlsx'
    wb.save(response)
    return response

#Modulo de ventas 
@login_required
def registrar_ventas (request):
    ok = False 
    errores = {}

    if request.method == 'POST':
        datos = request.POST
        form = VentasForm(datos)

        errores = validar_campos_especificos(post_data=datos)

        for campo, mensaje in errores.items():
            if campo in form.fields:
                form.add_error(campo, mensaje)

        if form.is_valid():
            venta = form.save(commit=False) 

            venta.id_admin = Administrador.objects.get(correo=request.user.username)
            
            producto = venta.id_producto

            if producto:
                precio_unitario = producto.precio_unitario
                total = venta.cantidad * precio_unitario

                if producto.stock >= venta.cantidad:
                    producto.stock -= venta.cantidad
                    producto.save()

                    venta.nombre_producto = producto.nombre
                    venta.precio_unitario = producto.precio_unitario
                    venta.total = total 
                    venta.contenido = producto.contenido
                    venta.unidad = producto.unidad

                    venta.save()

                    # Registro automático en ingresos
                    IngresosEgresos.objects.create(
                        id_admin=venta.id_admin,
                        tipo="Ingreso",
                        descripcion=f"Venta de {venta.cantidad} bolsas de {producto.nombre} de {producto.contenido} {producto.unidad} ",
                        fecha=venta.fecha,
                        monto=total
                    )

                    ok = True
                else:
                    form.add_error('cantidad', 'No hay suficiente stock disponible.')
            else:
                form.add_error('id_producto', 'Debes seleccionar un producto válido.')
    else:
        form = VentasForm()

    productos = Inventario.objects.filter(tipo="Inventario Producto final")
    return render(request, 'ingresos_egresos/registrar_ventas.html', {'form': form,'productos': productos,'ok': ok})

@login_required
def ventas (request):
    ok = False
    if request.method == "POST":

        seleccion = request.POST.get("elemento")       # columna seleccionada
        accion = request.POST.get("accion")         # accion enviada "editar" o "borrar"

        if seleccion:
            if "," in seleccion:
                ids = seleccion.split(",")  
            else:
                ids = [seleccion]
        else:
            ids = []

        ids = [int(x) for x in ids if x.strip().isdigit()]

        if accion == "borrar" and ids:
            ok = True
            for id_venta in ids:
                venta = get_object_or_404(Ventas, pk=id_venta)
                venta.anulada = True
                venta.save()

        elif accion == "editar" and len(ids) == 1:
            ok = True
            return redirect("actualizar_ventas", seleccion=ids[0])

    ventas = Ventas.objects.filter(anulada=False)
    cantidad_filas_vacias = 15 - ventas.count()

    contexto = {
        "ventas": ventas,
        "filas_vacias": range(cantidad_filas_vacias),
        "ok": ok,
    }
    return render (request, 'ingresos_egresos/mostrar_ventas.html', contexto)

# filtros y vistas para ventas
def filtrar_ventas(request):
    documento_admin = request.GET.get("documento_admin", "")
    producto = request.GET.get("producto", "")
    fecha = request.GET.get("fecha", "")
    cantidad = request.GET.get("cantidad", "")
    buscar = request.GET.get("buscar", "").strip()

    ventas = Ventas.objects.all()

    # Búsqueda general
    if buscar:
        buscar_normalizado = normalizar_texto(buscar)
        filtro_numerico = Q()
        filtro_texto = Q()
        filtro_fecha = Q()

        if es_numero(buscar_normalizado):
            try:
                filtro_numerico = (
                    Q(id_venta__iexact=buscar_normalizado) |
                    Q(id_admin__id_admin__iexact=buscar_normalizado) |
                    Q(cantidad=float(buscar_normalizado))
                )
            except (ValueError, InvalidOperation):
                filtro_numerico = Q()

        filtro_texto = (
            Q(id_producto__nombre__icontains=buscar_normalizado) |
            Q(id_admin__nombre__iexact=buscar_normalizado)
        )

        fecha_parseada = parsear_fecha(buscar_normalizado)
        if fecha_parseada:
            filtro_fecha = Q(fecha=fecha_parseada)

        ventas = ventas.filter(
            filtro_numerico | filtro_texto | filtro_fecha
        )

    if documento_admin:
        try:
            documento_admin_int = int(documento_admin)
            ventas = ventas.filter(id_admin__id_admin=documento_admin_int)
        except ValueError:
            pass

    if producto:
        ventas = ventas.filter(id_producto__nombre__icontains=producto)

    if fecha:
        ventas = ventas.filter(fecha=fecha)

    if cantidad:
        ventas = ventas.filter(cantidad=cantidad)

    return buscar, ventas, documento_admin, producto, fecha, cantidad

def actualizar_ventas(request, seleccion):
    ventas = get_object_or_404(Ventas, pk=seleccion)
    productos = Inventario.objects.filter(tipo="Inventario Producto final")

    if request.method == 'POST':
        id_producto = request.POST.get("id_producto", "").strip()
        cantidad = request.POST.get("cantidad", "").strip()

        # Validar que al menos un campo tenga valor
        if not id_producto and not cantidad:
            messages.error(request, "Debes ingresar al menos un dato para actualizar.")
            return redirect("actualizar_ventas", seleccion=seleccion)

        if id_producto:
            ventas.id_producto_id = id_producto  # Asignar FK por id

        if cantidad:
            ventas.cantidad = int(cantidad)  # Sin try, asumiendo valor válido

        ventas.save()
        return redirect('ventas')  

    return render(request, 'ingresos_egresos/actualizar_ventas.html', {
        'productos': productos,
        'ventas':ventas,
        })

@login_required
def informe_ventas (request):
    return render (request, 'ingresos_egresos/informe_ventas.html')

def datos_informe_ventas(request):
    # Obtener parámetros de la URL
    mes_str = request.GET.get('mes')
    anio_str = request.GET.get('anio')

    # Validación de existencia
    if not mes_str or not anio_str:
        return JsonResponse({'error': 'Debe enviar los parámetros "mes" y "anio".'}, status=400)

    # Validación de tipo
    try:
        mes = int(mes_str)
        anio = int(anio_str)
    except ValueError:
        return JsonResponse({'error': 'Los valores de mes y año deben ser números enteros.'}, status=400)

    # Validación de rango
    current_year = datetime.now().year
    if not (1 <= mes <= 12):
        return JsonResponse({'error': 'El mes debe estar entre 1 y 12.'}, status=400)
    if not (1900 <= anio <= current_year):
        return JsonResponse({'error': f'El año debe estar entre 1900 y {current_year}.'}, status=400)

    # Consulta a la base de datos: ventas agrupadas por producto
    ventas = (
        Ventas.objects.filter(fecha__year=anio, fecha__month=mes)
        .values('id_producto__nombre')
        .annotate(total_vendido=Sum('cantidad'))
        .order_by('-total_vendido')
    )

    # Si no hay datos, devolver advertencia
    if not ventas:
        return JsonResponse({
            'warning': f'No se encontraron ventas para {mes}/{anio}.',
            'labels': [],
            'ventas': []
        })

    # Preparar datos para el gráfico
    labels = [v['id_producto__nombre'] for v in ventas]
    data = [float(v['total_vendido']) for v in ventas]
    mas_vendido = labels[0] if labels else None

    return JsonResponse({
        'labels': labels,
        'ventas': data,
        'mas_vendido': mas_vendido
    })



# --- Exportar datos del gráfico de ventas mensual a Excel (GET, requiere mes y año) ---
@login_required
def exportar_ventas_grafico_excel(request):
    mes_str = request.GET.get('mes')
    anio_str = request.GET.get('anio')
    try:
        mes = int(mes_str)
        anio = int(anio_str)
    except (TypeError, ValueError):
        return HttpResponse('Parámetros de mes o año inválidos.', status=400)
    current_year = datetime.now().year
    if not (1 <= mes <= 12) or not (1900 <= anio <= current_year):
        return HttpResponse('Mes o año fuera de rango.', status=400)

    ventas = (
        Ventas.objects.filter(fecha__year=anio, fecha__month=mes)
        .values('id_producto__nombre', 'id_producto__precio_unitario')
        .annotate(total_vendido=Sum('cantidad'))
        .order_by('-total_vendido')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ventas_{mes}_{anio}"
    bold_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    fill_header = PatternFill(start_color='4b7510', end_color='4b7510', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:D1')
    ws['A1'] = f"Informe de Ventas - {mes}/{anio}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center

    ws.append(['Producto', 'Cantidad Vendida', 'Precio Unitario', 'Total'])
    for cell in ws[2]:
        cell.font = bold_font
        cell.alignment = center
        cell.fill = fill_header
        cell.border = border

    for v in ventas:
        precio = float(v['id_producto__precio_unitario']) if v['id_producto__precio_unitario'] is not None else 0.0
        cantidad = float(v['total_vendido'])
        total = precio * cantidad
        ws.append([
            v['id_producto__nombre'],
            cantidad,
            precio,
            total
        ])
        for cell in ws[ws.max_row]:
            cell.alignment = center
            cell.border = border

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="ventas_grafico_{mes}_{anio}.xlsx"'
    wb.save(response)
    return response

# --- Exportar ventas filtradas (POST, igual que ingresos_egresos) ---
@login_required
def exportar_ventas_excel(request):
    # Usar POST para recibir los filtros
    if request.method == 'POST':
        documento_admin = request.POST.get('documento_admin', '')
        producto = request.POST.get('producto', '')
        fecha = request.POST.get('fecha', '')
        cantidad = request.POST.get('cantidad', '')
        buscar = request.POST.get('buscar', '').strip()
    else:
        documento_admin = request.GET.get('documento_admin', '')
        producto = request.GET.get('producto', '')
        fecha = request.GET.get('fecha', '')
        cantidad = request.GET.get('cantidad', '')
        buscar = request.GET.get('buscar', '').strip()

    # Filtrar ventas igual que en filtrar_ventas
    ventas = Ventas.objects.all()
    if buscar:
        buscar_normalizado = normalizar_texto(buscar)
        filtro_numerico = Q()
        filtro_texto = Q()
        filtro_fecha = Q()
        if es_numero(buscar_normalizado):
            try:
                filtro_numerico = (
                    Q(id_venta__iexact=buscar_normalizado) |
                    Q(id_admin__id_admin__iexact=buscar_normalizado) |
                    Q(cantidad=float(buscar_normalizado))
                )
            except (ValueError, InvalidOperation):
                filtro_numerico = Q()
        filtro_texto = (
            Q(id_producto__nombre__icontains=buscar_normalizado) |
            Q(id_admin__nombre__iexact=buscar_normalizado)
        )
        fecha_parseada = parsear_fecha(buscar_normalizado)
        if fecha_parseada:
            filtro_fecha = Q(fecha=fecha_parseada)
        ventas = ventas.filter(filtro_numerico | filtro_texto | filtro_fecha)
    if documento_admin:
        try:
            documento_admin_int = int(documento_admin)
            ventas = ventas.filter(id_admin__id_admin=documento_admin_int)
        except ValueError:
            pass
    if producto:
        ventas = ventas.filter(id_producto__nombre__icontains=producto)
    if fecha:
        ventas = ventas.filter(fecha=fecha)
    if cantidad:
        ventas = ventas.filter(cantidad=cantidad)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    titulo_font = Font(bold=True, size=14)
    encabezado_font = Font(bold=True, color='FFFFFF')
    encabezado_fill = PatternFill(start_color='3D7A1F', end_color='3D7A1F', fill_type='solid')
    alineacion = Alignment(horizontal='center', vertical='center')
    borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fila_inicio = 2
    ws.merge_cells(start_row=fila_inicio, start_column=2, end_row=fila_inicio, end_column=8)
    celda_titulo = ws.cell(row=fila_inicio, column=2)
    celda_titulo.value = "Informe de Ventas"
    celda_titulo.font = titulo_font
    celda_titulo.alignment = alineacion
    encabezados = ['ID Venta', 'Documento Admin', 'Nombre Admin', 'Producto', 'Cantidad', 'Precio Unitario', 'Total', 'Fecha']
    for i, encabezado in enumerate(encabezados):
        col = 2 + i
        celda = ws.cell(row=fila_inicio + 2, column=col, value=encabezado)
        celda.font = encabezado_font
        celda.fill = encabezado_fill
        celda.alignment = alineacion
        celda.border = borde
    for index, item in enumerate(ventas):
        fila_actual = fila_inicio + 3 + index
        valores = [
            item.id_venta,
            item.id_admin.id_admin,
            f"{item.id_admin.nombre} {getattr(item.id_admin, 'apellido', '')}",
            f"{item.id_producto.nombre} {item.id_producto.contenido} {item.id_producto.unidad}",
            float(item.cantidad),
            float(getattr(item.id_producto, 'precio_unitario', 0)),
            float(item.total),
            item.fecha.strftime('%Y-%m-%d')
        ]
        for i, valor in enumerate(valores):
            celda = ws.cell(row=fila_actual, column=2 + i, value=valor)
            celda.alignment = alineacion
            celda.border = borde
    for col_letra in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col_letra].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ventas.xlsx'
    wb.save(response)
    return response

# --- Exportar ventas a PDF ---
@login_required
def exportar_ventas_pdf(request):
    _, ventas, documento_admin, producto, fecha, cantidad = filtrar_ventas(request)
    template_path = 'ingresos_egresos/pdf_ventas.html'
    context = {'ventas': ventas}
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ventas.pdf"'
    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response

